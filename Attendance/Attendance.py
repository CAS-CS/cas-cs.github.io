import numpy as np
import pandas as pd
import openpyxl as xl
# from openpyxl import workbook work
import flet as ft
from datetime import datetime
from dateutil.relativedelta import relativedelta 
import shutil
import sys
import os

import openpyxl as opl
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter
from openpyxl.styles import Font, Border, Color, Alignment, Side, PatternFill
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
BorderSide = Side(style='thin', color='FF000000')
BorderAll = Border(left=BorderSide, right=BorderSide,
                   top=BorderSide, bottom=BorderSide)
DateFont = Font(name='Calibri', size=10, bold=False, italic=False,
                underline='none', strike=False, color='FF000000')
DateAlignment = Alignment(horizontal='center', vertical='center',
                          text_rotation=90, wrap_text=False, shrink_to_fit=False, indent=0)

CellAlignment = Alignment(horizontal='center', vertical='center')

AbsentFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')

dxf = DifferentialStyle(font=Font(bold=True), fill=PatternFill(start_color='EE1111', end_color='11EE11'))
rule = Rule(type='cellIs', dxf=dxf, formula=["10"])


file = sys.argv[0]
print(file)
today = datetime.today().strftime('%Y-%m-%d')
# Read in the data
ThisFilePath = os.path.dirname(sys.argv[0])

print(os.path.dirname(sys.argv[0]))

fileName = "Attendance.xlsx"
filePath = os.path.join(ThisFilePath, fileName)
df = pd.read_excel(filePath)

attendance={}
def initAttendance(Names):
    attendance={}
    for name in Names:
        attendance[name] = dict(present=1)
    return attendance


def main(page: ft.Page):
##    attendance = read_names(fileName)

    # openpyxl functions

    def getHeaders(*args):
        print('getHeaders', args)
        fileName = args[0]

        wb = load_workbook(fileName)
        if len(args) > 1:
            sheetName = args[1]
        else:
            sheetName = wb.sheetnames[0]

        ws = wb[sheetName]

        cellRange = ws.dimensions
        cellRange
        R = range_boundaries(cellRange)
##        print(f"cell range{cellRange}")
##        print(f"sheet range {R}")
        startCol = R[0]
        endCol = R[2]
        startRow = R[1]
        endRow = R[3]
        
        print(f"Col Range={range(startCol,endCol)}")
        print(f"Row Range={range(startRow,endRow)}")

        # print(startCol,endCol,startRow,endRow)
        for col in range(startCol, endCol+1):
            for row in range(startRow, endRow+1):
##                print(f"{col} {row}")
##                print(ws.cell(row=row,column=col).value)
                if ws.cell(row=row, column=col).value == "Name":
                    HeaderRow = row
                    NameCol = col
                    break

        Names = []
        for row in range(startRow+1, endRow+1):
            testVal = str(ws.cell(row=row, column=NameCol).value)
            print(testVal)
            print(row)
            if testVal.lower().__contains__("class") | testVal.lower().__contains__("total") | testVal.lower().__contains__("none"):
                break
            elif len(testVal)==0:
                break
            elif testVal==" ":
                break
            Names.append(testVal)
        print(f'returning,Names={Names},HeaderRow={HeaderRow},NameCol={NameCol},endCol={endCol}')
        wb.close()
        return Names, HeaderRow, NameCol, endCol

    def appendColumn(*args):
        # appendColumn(fileName,sheetName,Header):
##        print(f"appendCOlumn", args)
        fileName = args[0]

        wb = load_workbook(fileName)
        if len(args) > 2:
            sheetName = args[1]
##            data = args[len(args)-1]
        else:
            sheetName = wb.sheetnames[0]
##            data = args[len(args)-1]
        ws = wb[sheetName]
        print(sheetName)
        Names, HeaderRow, NameCol, endCol = getHeaders(fileName, sheetName)
        
        dataVector = []
        Header = [str(HeaderUI.value)]
        AttendanceInput=getAttendanceFromInputUI()
        classNo=[str(endCol-NameCol+1)]
        Remark = [str(RemarkUI.value)]
        dataVector.extend(Header)
        dataVector.extend(AttendanceInput)
        dataVector.extend(classNo)
        dataVector.extend(Remark)


        
        data=dataVector
        


        #insert class number
        
        InsertColumn = []
        newCol = endCol+1
        ws.column_dimensions[get_column_letter(newCol)].width = 4
        

        for i, val in enumerate(data):
##            print(i, val)
            currentCell = ws.cell(row=HeaderRow+i, column=newCol)
            ws.row_dimensions[HeaderRow+i].height = 20
            if i == 0:
                currentCell.font = DateFont
                currentCell.alignment = DateAlignment
            elif i == len(data)-2:
                currentCell.alignment = DateAlignment
            elif i == len(data)-1:
                currentCell.alignment = DateAlignment
                ws.row_dimensions[HeaderRow+i].height = 250
            else:
                currentCell.alignment=CellAlignment
            currentCell.value = val
            currentCell.border = BorderAll
        ws.row_dimensions[HeaderRow].height = 90
        wb.save(fileName)
        wb.close()

    def getAttendanceSequence():
        seq = ''
        for i, name in enumerate(attendance):
            val = attendance[name]['present']
            breakSequence = 'meta,class,total,break'.lower().split(",")
            breakout = False
            for b in breakSequence:
                if str(name).lower().__contains__(b):
                    breakout = True
            if breakout == True:
                break
            if i % 4 == 0:
                seq += " "
            if val == 1:
                seq += "1"
            else:
                seq += "0"
        return seq

    def serialAttendanceToList(serialAttendance):
        listAllendance = serialAttendance.replace(" ", "").replace(",", "").replace(
            ";", "").replace("\n", "").replace("\t", "").replace("p", "1").replace("a", "0")
        listAllendance = list(listAllendance)
        return listAllendance


    def getAttendanceFromInputUI():
        listAllendance=serialAttendanceUI.value
        listAllendance = listAllendance.replace(" ", "").replace(",", "").replace(
            ";", "").replace("\n", "").replace("\t", "").replace("p", "1").replace("a", "0")
        listAllendance = list(listAllendance)
        return listAllendance

    def saveAttendanceFn(e):
        StatusUI.value = "Processing"
        StatusUI.update()
        serialAttendance = getAttendanceSequence()
        print(serialAttendance)
        saveAttendanceList = serialAttendanceToList(serialAttendance)
        # DF = pd.DataFrame({f"{saveDate}": saveAttendanceList})
        # DF = DF.reset_index(drop=True)
        # df = pd.read_excel(fileName).reset_index(drop=True)
        # df = pd.concat([df, DF], ignore_index=False, axis=1)
        # df = df.reset_index(drop=True)

        StatusUI.value = "Creating backup"
        StatusUI.update()
        shutil.copy(fileName, fileName+".bak")

        StatusUI.value = "Saving"
        StatusUI.update()
        # df.to_excel(fileName)
        appendColumn(fileName)
        os.system('xdg-open ' + fileName)

        StatusUI.value = "Saved"
        StatusUI.update()

        StatusUI.value = "Ready"
        StatusUI.update()

    def add_Button(target, label, data=True, index=""):
        def toggleAttendance(e):

            StatusUI.value = "Processing"
            StatusUI.update()
            if B.data["value"] == True:
                B.data["value"] = False
                B.bgcolor = ft.colors.RED_300
            else:
                B.data["value"] = True
                B.bgcolor = ft.colors.GREEN_300
            name = B.data['name']
            if B.data['value'] == False:
                present = 0
            else:
                present = 1
            attendance[name]['present'] = present
            print(f"Current State {B.data}")
            print(f"updated attendance {attendance[name]}")

            B.update()
            serialAttendance = getAttendanceSequence()
            serialAttendanceUI.value = serialAttendance
            serialAttendanceUI.update()

            StatusUI.value = "Ready"
            StatusUI.update()

        B = ft.FloatingActionButton(data=dict(name=label, value=True),
                                    shape=ft.RoundedRectangleBorder(radius=5),
                                    width=200,
                                    mini=True,
                                    bgcolor=ft.colors.GREEN_100,
                                    content=ft.Row(
                                        [ft.Text(f"{index} "),
                                         ft.Text(label,
                                                 size=14,
                                                 width=185,
                                                 selectable=False
                                                 )],
                                        alignment="center"),
                                    on_click=toggleAttendance)

        target.controls.append(B)

    print("starting main")
    page.title = "AttendanceUpdate"
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.window_height = 600
    page.window_width = 900

    Names, HeaderRow, NameCol, endCol = getHeaders(fileName)
    attendance=initAttendance(Names)
    page.scroll = ft.ScrollMode.ADAPTIVE
    MainRow = ft.Row()


    page.add(MainRow)

    MenuCol1 = ft.Column()
    MainRow.controls.append(MenuCol1)

    today=datetime.today()
    def minusHeader(e):
        today=today-relativedelta(days=1)
        HeaderUI.value=today
        HeaderUI.update()
    def plusHeader(e):
        today=today+relativedelta(days=1)
        HeaderUI.value=today
        HeaderUI.update()

    row0 = ft.Row()
##    row0.controls.append(ft.Dropdown(options=[
##            ft.dropdown.Option("Red"),
##            ft.dropdown.Option("Green"),
##            ft.dropdown.Option("Blue"),
##        ]))
    row1 = ft.Row()
    HeaderUI = ft.TextField(label="Date or Header", value=today.strftime('%Y-%m-%d'),width=200)
    HeaderUIMinus=ft.TextButton('-',on_click=minusHeader,width=40)
    HeaderUIPlus=ft.TextButton('+',on_click=plusHeader,width=40)
    row1.controls.append(HeaderUIMinus)
    row1.controls.append(HeaderUI)
    row1.controls.append(HeaderUIPlus)

    row2 = ft.Row()
    serialAttendanceUI = ft.TextField(label="Attendance")
    row2.controls.append(serialAttendanceUI)

    row3 = ft.Row()
    RemarkUI = ft.TextField(label="Remark", value="")
    row3.controls.append(RemarkUI)

    row4 = ft.Row()
    StatusUI = ft.Text(value="Ready")
    row4.controls.append(StatusUI)
    submitButton = ft.FloatingActionButton(
        content=ft.Row(
            [
                ft.Icon(ft.icons.ADD),
                ft.Text("Save Attendance")
            ],
            alignment="center", spacing=5
        ),
        bgcolor=ft.colors.BLUE_200,
        shape=ft.RoundedRectangleBorder(radius=5),
        mini=True,
        width=300,
        on_click=saveAttendanceFn
    )

    DefaultAttendanceRow = ft.Row()
    MenuCol1.controls.append(row0)
    MenuCol1.controls.append(row1)
    MenuCol1.controls.append(row2)
    MenuCol1.controls.append(row3)
    MenuCol1.controls.append(row4)
    MenuCol1.controls.append(DefaultAttendanceRow)
    MenuCol1.controls.append(submitButton)


    AttendanceToggleColumn2 = ft.Column()
    # AttendanceToggleColumn2.scroll = "AUTO"
    AttendanceToggleColumn2.wrap = True
    AttendanceToggleColumn2.height = page.window_height
    MainRow.controls.append(AttendanceToggleColumn2)

    for i, name in enumerate(attendance):
        val = attendance[name]['present']
        breakSequence = 'meta,class,total,break,remark,count'.lower().split(",")
        breakout = False
        for b in breakSequence:
            if str(name).lower().__contains__(b):
                breakout = True
        if breakout == True:
            break
        if val == 1:
            add_Button(AttendanceToggleColumn2, name, True, i+1)
        else:
            add_Button(AttendanceToggleColumn2, name, False, i+1)

    def page_resize(e):
        width = page.width
        height = page.height
        print(width, height)
        AttendanceToggleColumn2.height = height-80
        AttendanceToggleColumn2.update()

    page.on_resize = page_resize
    def web(e):
        print('web')
        os.system("xdg-open https://attendanceHelper.mgeek.in")
    page.add(ft.Row([ft.Text('Copyrights 2023 '),ft.TextButton("attendanceHelper.mgeek.in",on_click=web)]))
    page.update()

    
ft.app(target=main)
